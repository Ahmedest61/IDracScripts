IP_ADDRESS_COLUMN = 'A'
SERVICE_TAG_COLUMN = 'B'
MEMORY_COLUMN = 'C'
MEMORY_SLOT_COLUMN = 'D'
CPU_COLUMN = 'E'
CPU_THREAD_COLUMN = 'F'
NIC_COLUMN = 'G'
PORT_COLUMN = 'H'
DRIVE_COLUMN = 'I'
DRIVE_SIZE_COLUMN = 'J'
MEMORY_INFO_COLUMN = 'K'
CPU_INFO_COLUMN = 'L'
NIC_INFO_COLUMN = 'M'
HDD_DRIVE_INFO_COLUMN = 'N'
SSD_DRIVE_INFO_COLUMN = 'O'

USER_COLUMN = 'B'
PASSWORD_COLUMN = 'C'

try:
    import getpass
    import string
    import paramiko
    import re
    # import subprocess
    import math
    from math import *
    # from subprocess import *
    import os
    import json
  #  import pandas as pd
  #  from pandas.io.json import json_normalize
    from datetime import datetime
    from ConfigParser import *
    from ConfigParser import (ConfigParser, MissingSectionHeaderError,
                              ParsingError, DEFAULTSECT)
    import ConfigParser
    import collections
    from collections import OrderedDict, Counter
    from openpyxl import *
    from openpyxl.styles import *
except ImportError as error:
    print "You don't have module \"{0}\" installed.".format(error.message[16:])
    exit(1)


def createXlsxForIdrac(no_of_idracs, address, start):
    ip = address
    user = "root"
    password = "calvin"
    wb = Workbook()
    ws = wb.active
    ws.title = "IdracInfo"
    columns_heading = ["IP", "User", "Pass"]
    columns_alphabat = ['A', 'B', 'C']
    yellowFill = PatternFill(start_color='FFFFFF00',
                             end_color='FFFFFF00', fill_type='solid')
    i = 0
    while(i < len(columns_heading)):
        ws[columns_alphabat[i] + str(1)].fill = yellowFill
        ws[columns_alphabat[i] + str(1)] = columns_heading[i]
        i += 1
    index = 2

    for idrac in range(no_of_idracs):
        str_index = str(index)
        ws[IP_ADDRESS_COLUMN + str_index].value = ip + "." + str(start)
        ws[USER_COLUMN + str_index].value = user
        ws[PASSWORD_COLUMN + str_index].value = password
        start += 1
        index += 1

    name = 'IdracsList' + '.xlsx'
    wb.save(name)
    return name


def CreateXLS(filename):
    with open(filename) as data_file:
        data = json.load(data_file)
    wb = Workbook()
    ws = wb.active
    # Column names
    columns_heading = ["IP", "Service Tag", "Total Memory", "Memory Slots", "Total CPU", "Total threads", "Total Nics",
                       "Total Ports", "Total Drives", "Total Drives Size", "Memory Info", "CPU Info", "NIC Info", "HDD Drives Info", "SSD Drives Info"]
    # column names correspond to excel
    columns_alphabat = ['A', 'B', 'C', 'D', 'E',
                        'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    i = 0
    # color
    yellowFill = PatternFill(start_color='FFFFFF00',
                             end_color='FFFFFF00', fill_type='solid')
    # set column names with color
    while(i < len(columns_heading)):
        ws[columns_alphabat[i] + str(1)].fill = yellowFill
        ws[columns_alphabat[i] + str(1)] = columns_heading[i]
#		ws.column_dimensions[columns_alphabat[i]+str(1)].width = (len(columns_heading[i])+2)*1.2
        i += 1
    # Fill the data into the columns
    index = 2
    for idrac in range(len((data))):
        str_index = str(index)

        ws[IP_ADDRESS_COLUMN + str_index].value = data[idrac]['ip']
        ws[SERVICE_TAG_COLUMN + str_index].value = data[idrac]['service_tag']
        ws[MEMORY_COLUMN +
            str_index].value = data[idrac]['total_memory'].split("/")[1]
        ws[MEMORY_SLOT_COLUMN + str_index].value = data[idrac]['memory_slots']
        ws[CPU_COLUMN + str_index].value = data[idrac]['cpu_slots']
        ws[CPU_THREAD_COLUMN + str_index].value = data[idrac]['cpu_threads']
        ws[NIC_COLUMN + str_index].value = data[idrac]['nic_slots']
        ws[PORT_COLUMN + str_index].value = data[idrac]['ports']
        drives = []
        print data[idrac]['drives_type']
        # Get to knmow the types of drives i.e. HDD and SSD
        for ind in range(len(data[idrac]['drives_type'])):
            for key, value in data[idrac]['drives_type'].iteritems():
                if key not in drives:
                    drives.append(key)
        print drives
        tot_driv = 0
        # Calculate totaal drives using type (HDD or SSD)
        for ind in range(len(drives)):
           #         print data[idrac]['drives_type'][drives[ind]]
            tot_driv += int(data[idrac]['drives_type'][drives[ind]])
   #     print drives, tot_driv
        ws[DRIVE_COLUMN + str_index].value = tot_driv

        dumy_str = ''
        dumy_mem_model_list = []
        dumy_mem_speed_list = []
        mem_model_dict = dict()

        for mem_slot in range(ws[MEMORY_SLOT_COLUMN + str_index].value):
            dumy_mem_model_list.append(
                data[idrac]['memory_info'][mem_slot]['model'])
            dumy_mem_speed_list.append(
                data[idrac]['memory_info'][mem_slot]['speed'])
            if data[idrac]['memory_info'][mem_slot]['model'] not in mem_model_dict:
                mem_model_dict[data[idrac]['memory_info'][mem_slot]
                               ['model']] = data[idrac]['memory_info'][mem_slot]['speed']

        mem_model_set = Counter(dumy_mem_model_list)
  #      mem_model_set = collections.Counter(dumy_mem_model_list)
        print dumy_mem_model_list, mem_model_set, mem_model_dict
        for mem in range(len(mem_model_set)):
            dumy_str += str(mem_model_set.keys()[mem]) + " x " + "Count: " + str(mem_model_set.values()[
                mem]) + " Speed: " + mem_model_dict[str(mem_model_set.keys()[mem])] + "\n"
	    print dumy_str
        ws[MEMORY_INFO_COLUMN + str_index].value = dumy_str
#        for mem_slot in range(0, 1):
 #           dumy_str += str(ws[MEMORY_SLOT_COLUMN + str_index].value) + " x " + data[idrac]['memory_info'][mem_slot]['model'] + \
  #              " Speed: " + \
   #             data[idrac]['memory_info'][mem_slot]['speed'] + "\n"
    #    ws[MEMORY_INFO_COLUMN + str_index].value = dumy_str
#	print dumy_str
        dumy_str = ''
        for cpu_slot in range(ws[CPU_COLUMN + str_index].value):
            dumy_str += data[idrac]['cpu_info'][cpu_slot]['model'] + "\t" + \
                " Processors: " + \
                data[idrac]['cpu_info'][cpu_slot]['no_of_processors'] + "\n"
        ws[CPU_INFO_COLUMN + str_index].value = dumy_str

        dumy_str = ''
        for nic_slot in range(ws[NIC_COLUMN + str_index].value):
            #            for nested_index in range(len(data[idrac]['nic_info'][data[idrac]['nic_slots_name'][nic_slot]])):
            dumy_str += data[idrac]['nic_slots_name'][nic_slot] + \
                "\t" + " Type: " + data[idrac]['nic_info'][data[idrac]
                                                           ['nic_slots_name'][nic_slot]][0]['decription'] + "\n"
        ws[NIC_INFO_COLUMN + str_index].value = dumy_str

        typ = 'GB'
        driv_type_dict = dict()

        drives_size_list = [list()
                            for x in range(len(data[idrac]['drives_type']))]
        drives_info_list = [list()
                            for x in range(len(data[idrac]['drives_type']))]

#	for ind in range(len(data[idrac]['drives_type'])):
 #           drives.append(data[idrac]['drives_type'][ind])
        total_mem = 0
        # Using drives type (HDD/SSD), find and count common drives i.e. 800 GB x 4 drives and add them into the excel sheet
        for driv_type in range(len(data[idrac]['drives_type'])):
            dumy_str = ''
            driv_type_dict = {}
            # Adjust the size of hardrives i.e write 586 GB as 600, 1024 GB as 1TB and make a dictionary size vs number (600 GB x 4)
#	    print "Slots"
#	    print data[idrac]['drives_type'][drives[driv_type]]
      #      print data[idrac]['drives_type'][drives[driv_type]]
            for driv_slot in range(data[idrac]['drives_type'][drives[driv_type]]):
                size_val = float(data[idrac]['drives_info'][drives[driv_type]]
                                 [driv_slot]['total_size_in_giga_bytes'].split(" ")[0])
                if 150 < size_val < 200:
                    size_val = 200
                    typ = " GB"
                elif 200 < size_val < 250:
                    size_val = 250
                    typ = " GB"
                elif 250 < size_val < 300:
                    size_val = 300
                    typ = " GB"
                elif 400 < size_val < 500:
                    size_val = 500
                    typ = " GB"
                elif 500 < size_val < 600:
                    size_val = 600
                    typ = " GB"
                elif size_val > 1024:
                    size_val = round(size_val / float(1024), 2)
                    typ = " TB"
                else:
                    size_val = size_val
                    typ = " GB"
               # print size_val

                drives_size_list[driv_type].append(size_val)
                if str(size_val) not in driv_type_dict:
                    driv_type_dict[str(size_val)] = typ
            # Count the common drives
            drive_set = collections.Counter(drives_size_list[driv_type])
            print drive_set, driv_type_dict
            for drive in range(len(drive_set)):
                if "TB" in driv_type_dict[str(drive_set.keys()[drive])]:
                    drive_set.keys()[drive] = round(drive_set.keys()[drive], 2)
                    total_mem += (drive_set.keys()
                                  [drive]) * (drive_set.values()[drive]) * 1024
                elif "GB" in driv_type_dict[str(drive_set.keys()[drive])]:
                    # print type(drive_set.keys()[drive]), type(drive_set.values()[drive]), total_mem
                    total_mem += (drive_set.keys()
                                  [drive]) * (drive_set.values()[drive])
                dumy_str += str(drive_set.keys()[drive]) + driv_type_dict[str(drive_set.keys()[
                    drive])] + " x " + "Count: " + str(drive_set.values()[drive]) + "\n"
            # Save Drives Size vs No of drives
            drives_info_list[driv_type] = dumy_str
        print total_mem
        if total_mem > 1024:
            total_mem /= float(1024)
            total_mem = round(total_mem)
            total_mem = str(total_mem) + " TB"
        else:
            total_mem = str(total_mem) + " GB"
   #     print drives_info_list, total_mem
        for ind in range(len(drives)):
            if "HDD" in drives[ind]:
                ws[HDD_DRIVE_INFO_COLUMN + str_index].value = drives_info_list[ind]
            elif "SSD" in drives[ind]:
                ws[SSD_DRIVE_INFO_COLUMN + str_index].value = drives_info_list[ind]

        ws[DRIVE_SIZE_COLUMN +
            str_index].value = (total_mem)
        index += 1
    time = datetime.utcnow().strftime("%s")
    name = 'Output' + time + '.xlsx'
    wb.save(name)


def sshAndStoreData(IdracList):
    HwInventoryList = []
    for idracNo in range(0, len(IdracList)):

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            ssh.connect(IdracList[idracNo][0], username=IdracList[idracNo]
                        [1], password=IdracList[idracNo][2])
            stdin, stdout, stderr = ssh.exec_command("racadm hwinventory")
            HwInventoryList.append(stdout.readlines())
            ssh.close()
        except Exception as e:
            print "ssh.connect not work for Idrac with IP: ", IdracList[idracNo][0]
    return HwInventoryList


def main():
    while(1):
        type_of_input = raw_input(
            'Want to read Idrac info from Xls, press Y/y else any key for user defined idrac info: ')
        if type_of_input.isdigit():
            print "Enter a character or Enter E to exit"
            continue
        elif "E" in type_of_input:
            print "Program terminated successfully"
            exit()
        elif (type_of_input == "Y" or type_of_input == "y"):
            NoOfIdracs = raw_input('Enter the number of IDRACs :')
            NetworkAddress = raw_input(
                'Enter the first three byte of Network:')
            StartAddressLastByte = raw_input(
                'Enter the Last byte to start from :')
            dumy_ip = NetworkAddress + "." + StartAddressLastByte
            pat = re.compile("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")
            test = pat.match(dumy_ip)
            if test:
                pass
            else:
                print "Not a valid IP address to start from!"
                continue
            NoOfIdracs = int(NoOfIdracs)
            IdracListXlsxFile = createXlsxForIdrac(
                NoOfIdracs, str(NetworkAddress), int(StartAddressLastByte))
            miss_nxt_user_entry = 1
            break
    IdracList = []
    Idrac = []
    if miss_nxt_user_entry == 0:
        while(1):
            NoOfIdracs = raw_input('Enter the number of IDRACs :')
            if NoOfIdracs.isdigit():
                break
            elif "E" in NoOfIdracs:
                print "Program terminated successfully"
                exit()
            else:
                print "Enter a number or Enter E to exit"
                continue
        NoOfIdracs = int(NoOfIdracs)
        print NoOfIdracs
        # Store IP, User and Password into a list

        for i in range(0, NoOfIdracs):
            Idrac = []
            print "Idrac :", i
            while(1):
                pat = re.compile("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")
                IpAddress = raw_input('Enter the IP Address :')
                test = pat.match(IpAddress)
                if test:
                    pass
                else:
                    print "Please Enter a valid IP!"
                    continue
                Idrac.append(IpAddress)
                User = raw_input('Enter the user name :')
                Idrac.append(User)
            #	Pass = raw_input('Enter the Password :')
                Pass = getpass.getpass("Enter the password: ")
            #	print Pass
                Idrac.append(Pass)

                check = raw_input(
                    'Data is stored successfully, press N/n to re-enter or press any key to move forward :')
                if check == "N" or check == "n":
                    Idrac = []
                    continue
                else:
                    IdracList.append(Idrac)
                    print IdracList
                    break
    else:
        print "loading files"
        wb = load_workbook(IdracListXlsxFile)
        ws = wb['IdracInfo']
        xls_ind = 2

        for i in range(NoOfIdracs):
            Idrac = []
   #	    print IP_ADDRESS_COLUMN + str(xls_ind), ws[IP_ADDRESS_COLUMN + str(xls_ind)].value
            Idrac.append(ws[IP_ADDRESS_COLUMN + str(xls_ind)].value)
            Idrac.append(ws[USER_COLUMN + str(xls_ind)].value)
            Idrac.append(ws[PASSWORD_COLUMN + str(xls_ind)].value)
            IdracList.append(Idrac)
            xls_ind += 1
        print IdracList

    if len(IdracList) == 0:
        print "IdracList is empty, terminating the application"
        exit()
    # Using the stored information ssh into the server and get Hwinventory

    HwInventoryList = sshAndStoreData(IdracList)

    TotalIdracsOutputs = len(HwInventoryList)
    # Final list of idracs
    iDRAC_list = []

    Outfilename = "Idrac" + ".json"

    for IdracNo in range(0, TotalIdracsOutputs):

        # Lists and dictionaries variable initialization
        MemoryDictionary = {}
        idrac_dumy_dic = OrderedDict()

        Mem_section_list, mem_size_list, mem_speed_list, mem_descp_list, mem_serial_list, mem_model_list, mem_status_list = ([
        ]for i in range(0, 7))
        dumy_mem_dic = collections.defaultdict(dict)

        time = datetime.utcnow().strftime("%s")
        filename = "Idrac" + str(IdracNo) + ".ini"
        fptr = open(filename, "w+")

        # Read Hwinventory and write it into the file
        for line in HwInventoryList[IdracNo]:
            if "--" in line:
                continue
            fptr.write("%s\n" % line)
        fptr.close()
        # Read Sections from filename
        try:
            Config = ConfigParser.ConfigParser()
            Config.read(filename)
        except (ConfigParser.MissingSectionHeaderError,
                ConfigParser.ParsingError):
            print "No Section header"
        if " " in Config.sections():
            print "No section, exiting"
        # Store memory sections into the list
        Mem_section_list = [s for s, s in enumerate(
            Config.sections()) if "InstanceID: DIMM.Socket" in s]
        # Extract key, values from memory sections
        for mem_sec in Mem_section_list:
            mem_size_list += [Config.get(str(mem_sec), 'Size')]
            mem_speed_list += [Config.get(str(mem_sec), 'CurrentOperatingSpeed')]
            mem_descp_list += [Config.get(str(mem_sec), 'Manufacturer')]
            mem_serial_list += [Config.get(str(mem_sec), 'SerialNumber')]
            mem_model_list += [Config.get(str(mem_sec), 'Model')]
            mem_status_list += [Config.get(str(mem_sec), 'PrimaryStatus')]
        # Check the types and sizes of memory, convert it into MB, and store them into a list
        for index in range(0, len(mem_size_list)):
            if "MB" in mem_size_list[index]:
                value = mem_size_list[index].replace("MB", "")
                mem_size_list[index] = value
            elif "PB" in mem_size_list[index]:
                value = mem_size_list[index].replace("PB", "")
                value *= math.pow(1024, 3)
                mem_size_list[index] = value
            elif "TB" in mem_size_list[index]:
                value = mem_size_list[index].replace("TB", "")
                value *= math.pow(1024, 2)
                mem_size_list[index] = value
            elif "GB" in mem_size_list[index]:
                value = mem_size_list[index].replace("GB", "")
                value *= math.pow(1024, 1)
                mem_size_list[index] = value
            elif "KB" in mem_size_list[index]:
                value = mem_size_list[index].replace("KB", "")
                value /= 1024
                mem_size_list[index] = value

        mem_slots = len(mem_size_list)
        mem_info_dictionaries = [dict() for x in range(mem_slots)]
        mem_info_list = []

        # Store memory key value individually intothe list of dictionaries
        for i in range(mem_slots):
            mem_info_dictionaries[i]['manufacturer'] = mem_descp_list[i]
            mem_info_dictionaries[i]['speed'] = mem_speed_list[i]
            mem_info_dictionaries[i]['model'] = mem_model_list[i]
            mem_info_dictionaries[i]['serial'] = mem_serial_list[i]
            mem_info_dictionaries[i]['status'] = mem_status_list[i]
        # Store list of dictionaries into a list
        mem_info_list = mem_info_dictionaries

        idrac_dumy_dic['ip'] = str(IdracList[IdracNo][0])
        total_mem = str(sum([int(x) for x in mem_size_list]))
        idrac_dumy_dic['total_memory'] = total_mem + " MB" + \
            " / " + str(int(total_mem) / 1024) + " GB"
        idrac_dumy_dic['memory_slots'] = mem_slots
        idrac_dumy_dic['memory_info'] = mem_info_list

        # Lists variable initialization
        Cpu_section_list, cpu_processor_list, cpu_family_list, cpu_manufac_list, cpu_curr_clock_list, cpu_model_list = ([
        ]for i in range(6))
        cpu_prim_status_list, cpu_virt_list, cpu_voltag_list, cpu_enabled_thread_list, cpu_max_clock_speed_list = ([
        ]for i in range(5))
        cpu_ex_bus_clock_speed_list, cpu_hyper_thread_list, cpu_status_list = (
            []for i in range(3))

        # Store cpu sections into the list
        Cpu_section_list = [s for s, s in enumerate(
            Config.sections()) if "InstanceID: CPU.Socket" in s]
        # Extract key, values from cpu sections
        for cpu_sec in Cpu_section_list:
            cpu_processor_list += [Config.get(str(cpu_sec),
                                              'NumberOfProcessorCores')]
            cpu_family_list += [Config.get(str(cpu_sec), 'CPUFamily')]
            cpu_manufac_list += [Config.get(str(cpu_sec), 'Manufacturer')]
            cpu_curr_clock_list += [Config.get(str(cpu_sec),
                                               'CurrentClockSpeed')]
            cpu_model_list += [Config.get(str(cpu_sec), 'Model')]
            cpu_prim_status_list += [Config.get(str(cpu_sec), 'PrimaryStatus')]
            cpu_virt_list += [Config.get(str(cpu_sec),
                                         'VirtualizationTechnologyEnabled')]
            cpu_voltag_list += [Config.get(str(cpu_sec), 'Voltage')]
            cpu_enabled_thread_list += [Config.get(
                str(cpu_sec), 'NumberOfEnabledThreads')]
            cpu_max_clock_speed_list += [
                Config.get(str(cpu_sec), 'MaxClockSpeed')]
            cpu_ex_bus_clock_speed_list += [Config.get(
                str(cpu_sec), 'ExternalBusCLockSpeed')]
            cpu_hyper_thread_list += [Config.get(str(cpu_sec),
                                                 'HyperThreadingEnabled')]
            cpu_status_list += [Config.get(str(cpu_sec), 'CPUStatus')]

        cpu_slots = len(cpu_processor_list)
        cpu_info_dictionaries = [dict() for x in range(cpu_slots)]
        cpu_info_list = []
        total_threads = 0
        # Store cpu key value individually into the list of dictionaries
        for i in range(cpu_slots):
            cpu_info_dictionaries[i]['no_of_processors'] = cpu_processor_list[i]
            cpu_info_dictionaries[i]['cpu_family'] = cpu_family_list[i]
            cpu_info_dictionaries[i]['manufacturer'] = cpu_manufac_list[i]
            cpu_info_dictionaries[i]['current_clock_speed'] = cpu_curr_clock_list[i]
            cpu_info_dictionaries[i]['model'] = cpu_model_list[i]
            cpu_info_dictionaries[i]['primary_status'] = cpu_prim_status_list[i]
            cpu_info_dictionaries[i]['virtualization_technology_enabled'] = cpu_virt_list[i]
            cpu_info_dictionaries[i]['voltage'] = cpu_voltag_list[i]
            cpu_info_dictionaries[i]['no_of_enabled_thread'] = cpu_enabled_thread_list[i]
            cpu_info_dictionaries[i]['max_clock_Speed'] = cpu_max_clock_speed_list[i]
            cpu_info_dictionaries[i]['external_bus_clock_speed'] = cpu_ex_bus_clock_speed_list[i]
            cpu_info_dictionaries[i]['hyper_threading_enabled'] = cpu_hyper_thread_list[i]
            if cpu_hyper_thread_list[i] == "Yes" or cpu_hyper_thread_list[i] == "1":
                total_threads += int(cpu_enabled_thread_list[i])
            cpu_info_dictionaries[i]['cpu_status'] = cpu_status_list[i]
        # Store list of dictionaries into a list
        cpu_info_list = cpu_info_dictionaries
        idrac_dumy_dic['cpu_slots'] = cpu_slots
        idrac_dumy_dic['cpu_threads'] = total_threads

        idrac_dumy_dic['cpu_info'] = cpu_info_list
        # Lists variable initialization
        Nic_section_list, nic_manufac_list, nic_descrp_list, nic_dev_descrp_list, nic_fqdd_list = ([
        ]for i in range(5))
        nic_pci_sub_device_id_list, nic_pci_sub_vendor_id_list, nic_pci_device_id_list = (
            []for i in range(3))
        nic_pci_vendor_id_list, nic_bus_no_list, nic_slot_typ_list, nic_data_bus_width_list = ([
        ]for i in range(4))
        nic_fun_no_list, nic_dev_no_list = ([]for i in range(2))
        # Store nic sections into the list
        Nic_section_list = [s for s, s in enumerate(
            Config.sections()) if "InstanceID: NIC" in s]
        # Make a set of nics, this will be used for identifying the ports as a nic may contain more than 1 port
        Nic_set = set()
        port_slots = len(Nic_section_list)
        for section in Nic_section_list:
            nic = section.split(" ")[1].split(".")[0] + "." + section.split(" ")[1].split(
                ".")[1] + "." + section.split(" ")[1].split(".")[2].split("-")[0]
            if nic not in Nic_set:
                Nic_set.add(nic)
        nic_slots = len(Nic_set)
        # Using the set of nics, make a dictionary which corresponds to the ports each nic contains
        Nic_set_dic = dict()
        for nic in Nic_set:
            for nic_sec in Nic_section_list:
                if nic in nic_sec:
                    if nic not in Nic_set_dic:
                        Nic_set_dic[nic] = list()
                        Nic_set_dic[nic].append(nic_sec)
                    else:
                        Nic_set_dic[nic].append(nic_sec)

        # Extract key, values from nic sections
        for nic_sec in Nic_section_list:
            nic_manufac_list += [Config.get(str(nic_sec), 'Manufacturer')]
            nic_descrp_list += [Config.get(str(nic_sec), 'Description')]
            nic_dev_descrp_list += [Config.get(str(nic_sec),
                                               'DeviceDescription')]
            nic_fqdd_list += [Config.get(str(nic_sec), 'FQDD')]
            nic_pci_sub_device_id_list += [
                Config.get(str(nic_sec), 'PCISubDeviceID')]
            nic_pci_sub_vendor_id_list += [
                Config.get(str(nic_sec), 'PCISubVendorID')]
            nic_pci_device_id_list += [Config.get(str(nic_sec), 'PCIDeviceID')]
            nic_pci_vendor_id_list += [Config.get(str(nic_sec), 'PCIVendorID')]
            nic_bus_no_list += [Config.get(str(nic_sec), 'BusNumber')]
            nic_slot_typ_list += [Config.get(str(nic_sec), 'SlotType')]
            nic_data_bus_width_list += [
                Config.get(str(nic_sec), 'DataBusWidth')]
            nic_fun_no_list += [Config.get(str(nic_sec), 'FunctionNumber')]
            nic_dev_no_list += [Config.get(str(nic_sec), 'DeviceNumber')]

        nic_info_dictionaries = [dict() for x in range(port_slots)]
        # Store nic key value individually into the list of dictionaries
        for i in range(port_slots):
            nic_info_dictionaries[i]['manufacturer'] = nic_manufac_list[i]
            nic_info_dictionaries[i]['decription'] = nic_descrp_list[i]
            nic_info_dictionaries[i]['device_description'] = nic_dev_descrp_list[i]
            nic_info_dictionaries[i]['fqdd'] = nic_fqdd_list[i]
            nic_info_dictionaries[i]['pci_subdevice_id'] = nic_pci_sub_device_id_list[i]
            nic_info_dictionaries[i]['pci_Subvendor_id'] = nic_pci_sub_vendor_id_list[i]
            nic_info_dictionaries[i]['pci_Dev_id'] = nic_pci_device_id_list[i]
            nic_info_dictionaries[i]['pci_vendor_id'] = nic_pci_vendor_id_list[i]
            nic_info_dictionaries[i]['bus_number'] = nic_bus_no_list[i]
            nic_info_dictionaries[i]['slot_type'] = nic_slot_typ_list[i]
            nic_info_dictionaries[i]['data_bus_width'] = nic_data_bus_width_list[i]
            nic_info_dictionaries[i]['function_number'] = nic_fun_no_list[i]
            nic_info_dictionaries[i]['dev_number'] = nic_dev_no_list[i]
        # Store keys and values of nic in dictionary using a nic set vs port formate
        nic_port_dictionaries = dict()
        nic_name_list = []
        for nic in Nic_set:
            nic_name_list.append(nic)
            dumy_dict = [dict() for x in range(len(Nic_set_dic[nic]))]
            ind = 0
            for nics in Nic_set_dic[nic]:
                for j in range(port_slots):
                    if nic_info_dictionaries[j]['fqdd'] in nics:
                        dumy_dict[ind] = nic_info_dictionaries[j]
                        ind += 1
            nic_port_dictionaries[nic] = dumy_dict

        idrac_dumy_dic['nic_slots'] = nic_slots
        idrac_dumy_dic['nic_slots_name'] = nic_name_list
        idrac_dumy_dic['ports'] = port_slots
        idrac_dumy_dic['nic_info'] = nic_port_dictionaries
        # Lists variable initialization
        Drive_section_list, driv_dev_list, driv_dev_descrp_list, driv_ppid_list, temp_size_list = ([
        ]for i in range(5))
        driv_sas_address_list, driv_max_cap_speed_list, driv_used_siz_byte_list, driv_media_typ_list = ([
        ]for i in range(4))
        driv_blck_siz_byte_list, driv_bus_protocol_list, driv_serial_no_list, driv_manufacturer_list = ([
        ]for i in range(4))
        driv_fqdd_list, driv_slot_list, driv_raid_status_list, driv_predictiv_fail_status_list, driv_free_siz_byte_list, driv_total_siz_list = ([
        ]for i in range(6))
        # Store drive sections into the list
        Drive_section_list = [s for s, s in enumerate(
            Config.sections()) if ("InstanceID: Disk") in s]
        i = 0
        drive_list_len = len(Drive_section_list)
        # Filter the physical hardrives
        while(i < len(Drive_section_list)):
            if Config.has_option(Drive_section_list[i], "Device Type"):
                if "PhysicalDisk" not in Config.get(Drive_section_list[i], "Device Type"):
                    del Drive_section_list[i]
                else:
                    i += 1
                    continue
        # Extract key, values from drives sections
        check_ssd = 0
        check_hdd = 0
        for dev_sec in Drive_section_list:
            driv_dev_list += [Config.get(str(dev_sec), 'Device Type')]
            driv_dev_descrp_list += [Config.get(str(dev_sec),
                                                'DeviceDescription')]
            driv_ppid_list += [Config.get(str(dev_sec), 'PPID')]
            driv_sas_address_list += [Config.get(str(dev_sec), 'SASAddress')]
            driv_max_cap_speed_list += [
                Config.get(str(dev_sec), 'MaxCapableSpeed')]
            driv_used_siz_byte_list += [
                Config.get(str(dev_sec), 'UsedSizeInBytes')]
            driv_free_siz_byte_list += [
                Config.get(str(dev_sec), 'FreeSizeInBytes')]
            total_hard_size = int(Config.get(str(dev_sec), 'UsedSizeInBytes').split(" ")[0]) + int(
                Config.get(str(dev_sec), 'FreeSizeInBytes').split(" ")[0])
            total_hard_size /= math.pow(1024, 3)
#	    print total_hard_size
            driv_total_siz_list.append(total_hard_size)
            driv_media_typ_list += [Config.get(str(dev_sec), 'MediaType')]
            if ("Solid State Drive" in Config.get(str(dev_sec), 'MediaType')):
                check_ssd += 1
            if ("HDD" in Config.get(str(dev_sec), 'MediaType')):
                check_hdd += 1
            driv_blck_siz_byte_list += [
                Config.get(str(dev_sec), 'BlockSizeInBytes')]
            driv_bus_protocol_list += [Config.get(str(dev_sec), 'BusProtocol')]
            driv_serial_no_list += [Config.get(str(dev_sec), 'SerialNumber')]
            driv_manufacturer_list += [
                Config.get(str(dev_sec), 'Manufacturer')]
            driv_fqdd_list += [Config.get(str(dev_sec), 'FQDD')]
            driv_slot_list += [Config.get(str(dev_sec), 'Slot')]
            driv_raid_status_list += [Config.get(str(dev_sec), 'RaidStatus')]
            driv_predictiv_fail_status_list += [
                Config.get(str(dev_sec), 'PredictiveFailureState')]
        # Adjust the size of each hardrives
 #       for index in range(0, len(driv_total_byte_list)):
  #          value = int(driv_total_byte_list[index])
   #         value /= float(1024 * 1024 * 1024)
    #        driv_total_siz_list[index] = value
#        print driv_total_siz_list
        total_size = (sum([int(x) for x in driv_total_siz_list]))
        driv_slots = len(Drive_section_list)
        driv_info_hdd_dictionaries = [dict() for x in range(check_hdd)]
        driv_info_ssd_dictionaries = [dict() for x in range(check_ssd)]
        driv_info_list = []
        # Store hardrive key value individually into the list of dictionaries
        hdd_index = 0
        ssd_index = 0
        for i in range(driv_slots):
            if "HDD" in driv_media_typ_list[i]:
                driv_info_hdd_dictionaries[hdd_index]['device_type'] = driv_dev_list[i]
                driv_info_hdd_dictionaries[hdd_index]['device_description'] = driv_dev_descrp_list[i]
                driv_info_hdd_dictionaries[hdd_index]['ppid'] = driv_ppid_list[i]
                driv_info_hdd_dictionaries[hdd_index]['sas_address'] = driv_sas_address_list[i]
                driv_info_hdd_dictionaries[hdd_index]['max_capable_speed'] = driv_max_cap_speed_list[i]
                driv_info_hdd_dictionaries[hdd_index]['used_size_in_bytes'] = driv_used_siz_byte_list[i]
                driv_info_hdd_dictionaries[hdd_index]['free_size_in_bytes'] = driv_free_siz_byte_list[i]
                driv_info_hdd_dictionaries[hdd_index]['total_size_in_giga_bytes'] = str(
                    driv_total_siz_list[i]) + " GB"
#		    driv_info_hdd_dictionaries[hdd_index]['media_type'] = driv_media_typ_list[i]
                driv_info_hdd_dictionaries[hdd_index]['block_size_in_bytes'] = driv_blck_siz_byte_list[i]
                driv_info_hdd_dictionaries[hdd_index]['bus_protocol'] = driv_bus_protocol_list[i]
                driv_info_hdd_dictionaries[hdd_index]['serial_no'] = driv_serial_no_list[i]
                driv_info_hdd_dictionaries[hdd_index]['manufacturer'] = driv_manufacturer_list[i]
                driv_info_hdd_dictionaries[hdd_index]['fqdd'] = driv_fqdd_list[i]
                driv_info_hdd_dictionaries[hdd_index]['slot'] = driv_slot_list[i]
                driv_info_hdd_dictionaries[hdd_index]['raid_status'] = driv_raid_status_list[i]
                driv_info_hdd_dictionaries[hdd_index]['predictive_failure_state'] = driv_predictiv_fail_status_list[i]
                hdd_index += 1
            elif "Solid State Drive" in driv_media_typ_list[i]:
                driv_info_ssd_dictionaries[ssd_index]['device_type'] = driv_dev_list[i]
                driv_info_ssd_dictionaries[ssd_index]['device_description'] = driv_dev_descrp_list[i]
                driv_info_ssd_dictionaries[ssd_index]['ppid'] = driv_ppid_list[i]
                driv_info_ssd_dictionaries[ssd_index]['sas_address'] = driv_sas_address_list[i]
                driv_info_ssd_dictionaries[ssd_index]['max_capable_speed'] = driv_max_cap_speed_list[i]
                driv_info_ssd_dictionaries[ssd_index]['used_size_in_bytes'] = driv_used_siz_byte_list[i]
                driv_info_ssd_dictionaries[ssd_index]['free_size_in_bytes'] = driv_free_siz_byte_list[i]

                driv_info_ssd_dictionaries[ssd_index]['total_size_in_giga_bytes'] = str(
                    driv_total_siz_list[i]) + " GB"
#		    driv_info_ssd_dictionaries[ssd_index]['media_type'] = driv_media_typ_list[i]
                driv_info_ssd_dictionaries[ssd_index]['block_size_in_bytes'] = driv_blck_siz_byte_list[i]
                driv_info_ssd_dictionaries[ssd_index]['bus_protocol'] = driv_bus_protocol_list[i]
                driv_info_ssd_dictionaries[ssd_index]['serial_no'] = driv_serial_no_list[i]
                driv_info_ssd_dictionaries[ssd_index]['manufacturer'] = driv_manufacturer_list[i]
                driv_info_ssd_dictionaries[ssd_index]['fqdd'] = driv_fqdd_list[i]
                driv_info_ssd_dictionaries[ssd_index]['slot'] = driv_slot_list[i]
                driv_info_ssd_dictionaries[ssd_index]['raid_status'] = driv_raid_status_list[i]
                driv_info_ssd_dictionaries[ssd_index]['predictive_failure_state'] = driv_predictiv_fail_status_list[i]
                ssd_index += 1

        driv_dict = dict()
        driv_type_dict = dict()
#        driv_type_list = []
#        idrac_dumy_dic['drives'] = driv_slots
        if check_ssd > 0:
            driv_dict['SSD'] = driv_info_ssd_dictionaries
            driv_type_dict['SSD'] = ssd_index
        if check_hdd > 0:
            driv_dict['HDD'] = driv_info_hdd_dictionaries
            driv_type_dict['HDD'] = hdd_index
 #       driv_info_list.append(driv_dict)
#        driv_type_list.append(driv_type_dict)
        idrac_dumy_dic['drives_type'] = driv_type_dict
        idrac_dumy_dic['total_drives_size'] = str(total_size) + " GB"
        idrac_dumy_dic['drives_info'] = driv_dict

        # Store sys sections into the list
        Sys_section_list = [s for s, s in enumerate(
            Config.sections()) if "InstanceID: System.Embedded" in s]
        idrac_dumy_dic['service_tag'] = Config.get(
            str(Sys_section_list[0]), 'ServiceTag')

        # Finally add all the nested dictionaries and list into a list (json formate)
        iDRAC_list.append(idrac_dumy_dic)
    # Store the json formate into a file
    fptr = open(Outfilename, "w+")
    json.dump(iDRAC_list, fptr, indent=4,
              sort_keys=False, separators=(',', ':'))
    fptr.close()
    # Create Excel file from json
    CreateXLS(Outfilename)


main()
# df = pd.read_json(Outfilename, grab_nest=(cpu_info, drives_info, nic_info, memory_info))
# df = pd.read_json(Outfilename)
# json_normalize(df)
# print "writing to file"
# time=datetime.utcnow().strftime("%s")
# name='Output'+time+'.xlsx'
# df.to_excel(name)
